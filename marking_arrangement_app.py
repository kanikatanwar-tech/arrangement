import streamlit as st
import pandas as pd
import math
import re
import io
from datetime import date
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Marking Arrangement", layout="wide")
st.title("📋 Teacher Marking Arrangement")

PERIOD_COLS = ["P0","P1","P2","P3","P4","P5","P6","P7","P8"]
DAY_MAP = {
    "Monday":"MON","Tuesday":"TUE","Wednesday":"WED",
    "Thursday":"THU","Friday":"FRI","Saturday":"SAT"
}
FIRST_HALF  = PERIOD_COLS[:math.ceil(len(PERIOD_COLS)/2)]   # P0–P4 (5 periods)
SECOND_HALF = PERIOD_COLS[math.ceil(len(PERIOD_COLS)/2):]   # P5–P8 (4 periods)

# ═══════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════

def clean_name(raw: str) -> str:
    """
    Strip ID/designation suffix after the last hyphen when the suffix contains no spaces.
    Handles both numeric IDs and designation codes:
      'AAKRITI SACHDEVA-20232584' → 'AAKRITI SACHDEVA'
      'ANAMIKA TOMAR-GT'          → 'ANAMIKA TOMAR'
      'ANJU-PTVT'                 → 'ANJU'
      'NEHA TYAGI-GT'             → 'NEHA TYAGI'
      'NEW SOCIO-PGT'             → 'NEW SOCIO'
    A suffix is stripped only if it has no internal spaces (i.e. it is a code, not a name part).
    """
    s = raw.strip()
    m = re.match(r'^(.+)-([^\s-]+)\s*$', s)
    if m and ' ' not in m.group(2):          # suffix is a single token → strip it
        return m.group(1).strip()
    return s

def norm(name: str) -> str:
    """Normalise for fuzzy matching: uppercase, collapse spaces."""
    return re.sub(r'\s+', ' ', name.strip().upper())

def get_half_cols(leave_type: str) -> list:
    if leave_type == "Full Day":     return PERIOD_COLS[:]
    elif leave_type == "First Half": return FIRST_HALF[:]
    else:                            return SECOND_HALF[:]

def which_half(pc: str) -> str:
    return "First Half" if pc in FIRST_HALF else "Second Half"

def extract_teacher_from_class_cell(cell: str) -> list:
    """
    'SCI (AMBICA MALHOTRA)' → ['AMBICA MALHOTRA']
    'PJB (AMARPREET KAUR)\\nSKT (SONIA DEVI)' → ['AMARPREET KAUR', 'SONIA DEVI']
    """
    if not cell:
        return []
    result = []
    for m in re.findall(r'\(([^)]+)\)', cell):
        for part in m.split('/'):
            t = part.strip()
            if t and t.lower() != 'nan':
                result.append(t)
    return result

def get_classes_from_teacher_cell(cell: str) -> list:
    """
    Extract ALL class names from a Teacher TT cell.
    '8D HIN'        → ['8D']
    '12C - NEEEV'   → ['12C']
    '9B+9D PJB'     → ['9B', '9D']
    '8B+8D+8E PJB'  → ['8B', '8D', '8E']
    """
    if not cell: return []
    return list(dict.fromkeys(re.findall(r'\b(\d+[A-Z])\b', cell)))

# ═══════════════════════════════════════════
#  DATA LOADING
# ═══════════════════════════════════════════

@st.cache_data
def load_summary(file_bytes: bytes):
    """Load Teacher TT, Class TT, Unavailability from Summary_Report."""
    xl = pd.ExcelFile(io.BytesIO(file_bytes))

    # ── Teacher TT ──────────────────────────
    raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Teacher TT Summary", header=None)
    tt_rows, cur = [], None
    for _, row in raw.iloc[2:].iterrows():
        if pd.notna(row[0]) and str(row[0]).strip():
            cur = str(row[0]).strip()
        if pd.isna(row[1]) or cur is None:
            continue
        rec = {"Teacher_raw": cur,
               "Teacher": clean_name(cur),
               "Day": str(row[1]).strip()}
        for i, pc in enumerate(PERIOD_COLS):
            v = row[i+2] if i+2 < len(row) else None
            rec[pc] = str(v).strip() if pd.notna(v) and str(v).strip() not in ["nan",""] else ""
        tt_rows.append(rec)
    teacher_tt = pd.DataFrame(tt_rows)

    # ── Class TT ────────────────────────────
    # Build: { class: { day: { period: [teacher_display_names] } } }
    raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Class TT Summary", header=None)
    class_tt = {}   # plain dict — no lambdas so st.cache_data can pickle it
    cur_class = None
    for _, row in raw.iloc[2:].iterrows():
        if pd.notna(row[0]) and str(row[0]).strip():
            cur_class = str(row[0]).strip()
        if pd.isna(row[1]) or cur_class is None:
            continue
        day = str(row[1]).strip()
        for i, pc in enumerate(PERIOD_COLS):
            v = row[i+2] if i+2 < len(row) else None
            cell = str(v).strip() if pd.notna(v) and str(v).strip() not in ["nan",""] else ""
            teachers = extract_teacher_from_class_cell(cell)
            if cur_class not in class_tt:
                class_tt[cur_class] = {}
            if day not in class_tt[cur_class]:
                class_tt[cur_class][day] = {}
            class_tt[cur_class][day][pc] = teachers

    # ── Unavailability ───────────────────────
    # { teacher_display_name (normed) : set of period cols }
    unavail = {}
    if "Unavailability" in xl.sheet_names:
        raw_u = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Unavailability", header=None)
        for _, row in raw_u.iloc[1:].iterrows():
            t, p = row[0], row[1]
            if pd.isna(t) or pd.isna(p): continue
            t_disp = clean_name(str(t).strip())
            pset = set()
            for x in str(p).split(','):
                x = x.strip()
                if x.isdigit():
                    pset.add(f"P{x}")
            if pset:
                unavail[norm(t_disp)] = pset

    # ── Teacher display names ────────────────
    all_display = sorted(set(teacher_tt["Teacher"].tolist()))
    return teacher_tt, class_tt, unavail, all_display



@st.cache_data
def load_ct_coct(file_bytes: bytes) -> dict:
    """
    Load CT_CO-CT.xlsx.
    The file has full names with ID suffixes identical to Teacher TT Summary.
    Returns { class_name: { "CT": name, "CO-CT": name, "2nd CO-CT": name|None } }
    clean_name() strips the ID suffix so names match Day_wise_TT exactly.
    """
    raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
    result = {}
    for _, row in raw.iloc[1:].iterrows():
        cls = str(row[0]).strip() if pd.notna(row[0]) else ''
        if not cls or cls == 'nan':
            continue
        def _get(val):
            v = str(val).strip() if pd.notna(val) else ''
            return clean_name(v) if v and v.lower() != 'nan' else None
        result[cls] = {
            "CT"       : _get(row[1]),
            "CO-CT"    : _get(row[2]),
            "2nd CO-CT": _get(row[3]),
        }
    return result

@st.cache_data
def load_day_tt(file_bytes: bytes):
    """
    Load Day_wise_TT.xlsx.
    Returns: { day: { teacher_display_name: { pc: class_or_empty } } }
    Day_wise_TT is the ground truth for FREE/BUSY per teacher per period.
    """
    result = {}
    for day in ["MON","TUE","WED","THU","FRI","SAT"]:
        try:
            raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=day, header=None)
        except Exception:
            continue

        # Find header row
        header_row = 0
        for i, row in raw.iterrows():
            if str(row[0]).strip() in ["Teacher Name","TEACHER NAME"]:
                header_row = i
                break

        teachers = {}
        for _, row in raw.iloc[header_row+1:].iterrows():
            name = str(row[0]).strip()
            if not name or name.lower() == "nan":
                continue
            rec = {}
            for i, pc in enumerate(PERIOD_COLS):
                v = row[i+1] if i+1 < len(row) else None
                rec[pc] = str(v).strip() if pd.notna(v) and str(v).strip() not in ["nan",""] else ""
            teachers[name] = rec
        result[day] = teachers
    return result

# ═══════════════════════════════════════════
#  FREE / BUSY LOGIC  (Day_wise_TT is source of truth)
# ═══════════════════════════════════════════

def get_unavail_set(teacher_name: str, unavail: dict) -> set:
    """Return set of period cols the teacher is marked unavailable for."""
    key = norm(teacher_name)
    return unavail.get(key, set())

def count_busy_in_half(teacher_name: str, day: str,
                       half_cols: list,
                       day_tt: dict, unavail: dict,
                       extra_busy: dict) -> dict:
    """
    Count busy periods within half_cols using Day_wise_TT + unavailability + extra_busy.
    Returns { tt: int, unav: int, extra: int, total: int }
    """
    teacher_day = day_tt.get(day, {}).get(teacher_name, {})
    tt_busy = {pc for pc in half_cols if teacher_day.get(pc, "")}

    unav_set  = get_unavail_set(teacher_name, unavail)
    unav_busy = {pc for pc in half_cols if pc in unav_set}

    extra_set = set(extra_busy.get(norm(teacher_name), set())) & set(half_cols)

    total = tt_busy | unav_busy | extra_set
    return {"tt": len(tt_busy), "unav": len(unav_busy),
            "extra": len(extra_set), "total": len(total)}

def is_free_in_period(teacher_name: str, day: str, pc: str,
                      day_tt: dict, unavail: dict,
                      extra_busy: dict) -> bool:
    """True only if teacher has NO class in this period, not unavailable, not extra-assigned."""
    teacher_day = day_tt.get(day, {}).get(teacher_name, {})
    if teacher_day.get(pc, ""):
        return False
    if pc in get_unavail_set(teacher_name, unavail):
        return False
    if pc in extra_busy.get(norm(teacher_name), set()):
        return False
    return True

def get_all_same_class_teachers(teacher_name: str, pc: str, day: str,
                                teacher_tt, class_tt: dict) -> list:
    """
    Find ALL teachers who teach the SAME CLASS(ES) as 'teacher_name' in period 'pc',
    looking across EVERY day in the Class TT (not just the arrangement day).
    Rationale: a teacher who teaches class 7B on MON/WED is a valid same-class
    substitute for a 7B period on FRI even if not scheduled for 7B that day.
    Handles multi-class cells like '9B+9D PJB'.
    Returns list of unique teacher display names.
    """
    rows = teacher_tt[(teacher_tt["Teacher"] == teacher_name) & (teacher_tt["Day"] == day)]
    if rows.empty:
        return []
    cell_val = rows.iloc[0].get(pc, "")

    # Extract ALL classes from the cell (handles combined classes like 9B+9D)
    klasses = get_classes_from_teacher_cell(cell_val)
    if not klasses:
        return []

    # Union of all teachers across every class found, across ALL days in Class TT
    teachers = set()
    for klass in klasses:
        for day_data in class_tt.get(klass, {}).values():   # iterate all days, not just 'day'
            for period_teachers in day_data.values():
                for t in period_teachers:
                    teachers.add(t)
    return list(teachers)

# ═══════════════════════════════════════════
#  CANDIDATE LIST BUILDER
# ═══════════════════════════════════════════

def get_candidate_list(mode: str,
                       absent_teacher: str,   # display name
                       pc: str,
                       day: str,
                       teacher_tt,
                       class_tt: dict,
                       day_tt: dict,
                       unavail: dict,
                       all_display: list,
                       on_leave_names: set,   # display names of ALL teachers on leave today
                       extra_busy: dict,
                       ct_coct: dict = None) -> tuple:
    """
    Returns (candidates_list, fallback_used: bool, note: str).

    For P0 + Same Class Teacher: CO-CT appears first, then 2nd CO-CT, then remaining
    free same-class teachers, then fallback to all teachers if needed.
    Each candidate dict has an optional 'role' key for display in brackets.

    fallback_used: True when Same Class Teacher had no free candidates
                   and the pool was automatically widened to Any Teacher.
    note: human-readable string explaining any fallback.
    """
    exclude = on_leave_names | {absent_teacher}

    def _make_candidate(t, role=None):
        fh = count_busy_in_half(t, day, FIRST_HALF,  day_tt, unavail, extra_busy)
        sh = count_busy_in_half(t, day, SECOND_HALF, day_tt, unavail, extra_busy)
        return {
            "name"          : t,
            "role"          : role,      # e.g. "CO-CT" or "2nd CO-CT" or None
            "free_in_period": True,
            "fh"            : fh,
            "sh"            : sh,
            "fd_total"      : fh["total"] + sh["total"],
            "half_busy"     : (fh if pc in FIRST_HALF else sh)["total"],
        }

    def _build_free_list(raw_pool, role_map=None):
        """Build sorted free-teacher list. role_map: {teacher_name: role_label}"""
        pool = [t for t in raw_pool if t not in exclude]
        result = []
        for t in pool:
            if not is_free_in_period(t, day, pc, day_tt, unavail, extra_busy):
                continue
            role = role_map.get(t) if role_map else None
            result.append(_make_candidate(t, role))
        result.sort(key=lambda x: x["half_busy"])
        return result

    # ── P0 + Same Class Teacher: use CT_CO-CT priority ───────────
    if pc == "P0" and mode == "Same Class Teacher" and ct_coct:
        # Find the class the absent teacher is covering in P0
        rows = teacher_tt[(teacher_tt["Teacher"] == absent_teacher) & (teacher_tt["Day"] == day)]
        klass = None
        if not rows.empty:
            klasses = get_classes_from_teacher_cell(rows.iloc[0].get("P0", ""))
            klass = klasses[0] if klasses else None

        if klass and klass in ct_coct:
            entry   = ct_coct[klass]
            co_ct   = entry.get("CO-CT")        # full Day_wise_TT name or None
            snd_ct  = entry.get("2nd CO-CT")    # full Day_wise_TT name or None

            # Priority list: CO-CT first, then 2nd CO-CT — only if free
            priority = []
            seen = set()
            for name, role in [(co_ct, "CO-CT"), (snd_ct, "2nd CO-CT")]:
                if (name and name not in exclude and name not in seen
                        and is_free_in_period(name, day, pc, day_tt, unavail, extra_busy)):
                    priority.append(_make_candidate(name, role))
                    seen.add(name)

            # Rest of same-class teachers (free, not already in priority)
            same_pool = get_all_same_class_teachers(absent_teacher, pc, day, teacher_tt, class_tt)
            rest = _build_free_list(
                [t for t in same_pool if t not in seen],
                role_map={}
            )

            candidates = priority + rest
            if candidates:
                return candidates, False, ""

            # Fallback to all teachers if no same-class candidate free
            any_pool = list(day_tt.get(day, {}).keys())
            fallback = _build_free_list([t for t in any_pool if t not in seen])
            # Re-add priority candidates to front even from any-teacher pool
            for name, role in [(co_ct, "CO-CT"), (snd_ct, "2nd CO-CT")]:
                if (name and name not in exclude
                        and is_free_in_period(name, day, pc, day_tt, unavail, extra_busy)):
                    fallback_names = [c["name"] for c in fallback]
                    if name in fallback_names:
                        idx = fallback_names.index(name)
                        cand = fallback.pop(idx)
                        cand["role"] = role
                        fallback.insert(0, cand)
            return fallback, True, "No free same-class teacher — showing all free teachers"

    # ── Standard Same Class Teacher ───────────────────────────────
    if mode == "Same Class Teacher":
        same_class_pool = get_all_same_class_teachers(absent_teacher, pc, day, teacher_tt, class_tt)
        candidates = _build_free_list(same_class_pool)
        if candidates:
            return candidates, False, ""
        any_pool = list(day_tt.get(day, {}).keys())
        return _build_free_list(any_pool), True, "No free same-class teacher — showing all free teachers"

    else:  # Any Teacher
        any_pool = list(day_tt.get(day, {}).keys())
        return _build_free_list(any_pool), False, ""

# ═══════════════════════════════════════════
#  REMARKS STRING
# ═══════════════════════════════════════════

def remarks_str(candidate: dict, leave_type: str) -> str:
    fh = candidate["fh"]
    sh = candidate["sh"]
    fd = candidate["fd_total"]

    def fmt(info, n):
        parts = [f"TT:{info['tt']}"]
        if info["unav"]:  parts.append(f"Unavail:{info['unav']}")
        if info["extra"]: parts.append(f"Arranged:{info['extra']}")
        parts.append(f"**{info['total']}/{n}**")
        return " ".join(parts)

    if leave_type == "First Half":
        return fmt(fh, len(FIRST_HALF)) + " busy in 1st Half"
    elif leave_type == "Second Half":
        return fmt(sh, len(SECOND_HALF)) + " busy in 2nd Half"
    else:
        return (f"1st:{fh['total']}/{len(FIRST_HALF)}  "
                f"2nd:{sh['total']}/{len(SECOND_HALF)}  "
                f"Full:**{fd}/{len(PERIOD_COLS)}**"
                + (f"  Arranged:{fh['extra']+sh['extra']}" if fh['extra']+sh['extra'] else ""))

# ═══════════════════════════════════════════
#  EXCEL EXPORT
# ═══════════════════════════════════════════

def _write_sheet(ws, sel_date, day_code, arrangement_rows, arr_state):
    H_FONT  = Font(bold=True, size=13, color="FFFFFF")
    H_FILL  = PatternFill("solid", fgColor="1F3864")
    CH_FONT = Font(bold=True, size=11, color="FFFFFF")
    CH_FILL = PatternFill("solid", fgColor="2E75B6")
    L_FILL  = PatternFill("solid", fgColor="FCE4D6")
    OK_FILL = PatternFill("solid", fgColor="E2EFDA")
    PD_FILL = PatternFill("solid", fgColor="FFF2CC")
    NO_FILL = PatternFill("solid", fgColor="F2F2F2")
    C_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L_ALIGN = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin    = Side(style="thin")
    BDR     = Border(left=thin, right=thin, top=thin, bottom=thin)

    last_col = get_column_letter(1 + len(PERIOD_COLS))
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = (f"Arrangement Sheet  |  "
               f"Date: {sel_date.strftime('%d %B %Y')}  |  "
               f"Day: {sel_date.strftime('%A')}  ({day_code})")
    c.font=H_FONT; c.fill=H_FILL; c.alignment=C_ALIGN
    ws.row_dimensions[1].height = 26

    for ci, h in enumerate(["Teacher on Leave"] + PERIOD_COLS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font=CH_FONT; c.fill=CH_FILL; c.alignment=C_ALIGN; c.border=BDR
    ws.row_dimensions[2].height = 20

    # Build teacher → period map
    # Only write arranged_by when the row has been explicitly confirmed (saved=True)
    teacher_map = defaultdict(dict)
    for row, state in zip(arrangement_rows, arr_state):
        if not row["period_col"]: continue
        # arranged_by is blank unless user has clicked Confirm
        arranged_by = state.get("_locked_name") if state.get("saved") else ""
        teacher_map[row["teacher"]][row["period_col"]] = {
            "class_subject": row["class_subject"],
            "arranged_by"  : arranged_by or "",
            "saved"        : state["saved"],
        }

    data_row = 3
    for teacher_disp, period_data in teacher_map.items():
        c = ws.cell(row=data_row, column=1, value=teacher_disp)
        c.font=Font(bold=True); c.fill=L_FILL; c.alignment=L_ALIGN; c.border=BDR
        for ci, pc in enumerate(PERIOD_COLS, 2):
            c = ws.cell(row=data_row, column=ci)
            c.alignment=C_ALIGN; c.border=BDR
            if pc in period_data:
                info = period_data[pc]
                by   = info["arranged_by"]
                if info["saved"] and by:
                    # Confirmed — show class + arranged teacher, green fill
                    c.value = f"{info['class_subject']}\n→ {by}"
                    c.fill  = OK_FILL
                elif info["saved"]:
                    # Confirmed but no teacher name (edge case)
                    c.value = info["class_subject"]
                    c.fill  = OK_FILL
                else:
                    # Not yet confirmed — show only class subject, leave arranger blank
                    c.value = info["class_subject"]
                    c.fill  = PD_FILL
            else:
                c.value = "-"; c.fill = NO_FILL
        ws.row_dimensions[data_row].height = 38
        data_row += 1

    ws.column_dimensions["A"].width = 28
    for ci in range(2, 2 + len(PERIOD_COLS)):
        ws.column_dimensions[get_column_letter(ci)].width = 20


def build_standalone(sel_date, day_code, arrangement_rows, arr_state) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sel_date.strftime("%d-%m-%Y")
    _write_sheet(ws, sel_date, day_code, arrangement_rows, arr_state)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()


def build_appended(file_bytes, sel_date, day_code, arrangement_rows, arr_state) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    sn = sel_date.strftime("%d-%m-%Y")
    if sn in wb.sheetnames:
        del wb[sn]
    ws = wb.create_sheet(title=sn)
    _write_sheet(ws, sel_date, day_code, arrangement_rows, arr_state)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()

# ═══════════════════════════════════════════════════════════
#  UI — STEP 1: Upload both files
# ═══════════════════════════════════════════════════════════

st.header("Step 1: Upload Files")
col_a, col_b, col_c = st.columns(3)
with col_a:
    summary_file = st.file_uploader("📊 Summary_Report.xlsx", type=["xlsx"],
                                     key="summary_upload")
with col_b:
    daywisett_file = st.file_uploader("📅 Day_wise_TT.xlsx", type=["xlsx"],
                                      key="daytt_upload")
with col_c:
    ctcoct_file = st.file_uploader("📋 CT_CO-CT.xlsx  *(optional — for P0 priority)*",
                                   type=["xlsx"], key="ctcoct_upload")

if not summary_file or not daywisett_file:
    st.info("Please upload Summary_Report and Day_wise_TT files to continue.")
    st.stop()

summary_bytes  = summary_file.read()
daywisett_bytes= daywisett_file.read()

teacher_tt, class_tt, unavail, all_display = load_summary(summary_bytes)
day_tt = load_day_tt(daywisett_bytes)

# Load CT_CO-CT if provided
ct_coct = {}
if ctcoct_file:
    ctcoct_bytes = ctcoct_file.read()
    ct_coct = load_ct_coct(ctcoct_bytes)

has_unav = len(unavail) > 0
st.success(
    f"✅ Summary loaded — Teachers: **{len(all_display)}** · "
    f"Classes: **{len(class_tt)}** · "
    f"Unavailability: **{'Yes – ' + str(len(unavail)) + ' entries' if has_unav else 'Not found'}**  "
    f"  |  Day_wise_TT loaded — Days: **{list(day_tt.keys())}**"
    + (f"  |  CT/CO-CT: **{len(ct_coct)} classes**" if ct_coct else "  |  CT/CO-CT: *not uploaded*")
)
st.caption(
    f"Period split → First Half: {FIRST_HALF} ({len(FIRST_HALF)} periods)  |  "
    f"Second Half: {SECOND_HALF} ({len(SECOND_HALF)} periods)"
)

# ═══════════════════════════════════════════════════════════
#  UI — STEP 2: Date + Leave entries
# ═══════════════════════════════════════════════════════════

st.header("Step 2: Select Date & Add Teachers on Leave")

col_d, _ = st.columns([2, 5])
sel_date = col_d.date_input("Select Date", value=date.today())
weekday  = sel_date.strftime("%A")
day_code = DAY_MAP.get(weekday)
if not day_code:
    st.warning(f"{weekday} is not a school day.")
    st.stop()
if day_code not in day_tt:
    st.error(f"Day '{day_code}' not found in Day_wise_TT.xlsx. Available: {list(day_tt.keys())}")
    st.stop()
st.write(f"**Day:** {weekday} ({day_code})")

# Use Day_wise_TT teacher names as the canonical list for the leave dropdown
day_tt_names = sorted(day_tt[day_code].keys())

if "leave_list" not in st.session_state:
    st.session_state.leave_list = []

with st.expander("➕ Add Teacher on Leave", expanded=True):
    a, b, c = st.columns([3, 2, 1])
    sel_t = a.selectbox("Teacher", day_tt_names, key="inp_t")
    sel_l = b.selectbox("Leave Type", ["First Half","Second Half","Full Day"], key="inp_l")
    c.write(""); c.write("")
    if c.button("Add"):
        if any(e["display"] == sel_t for e in st.session_state.leave_list):
            st.warning(f"{sel_t} is already in the list.")
        else:
            st.session_state.leave_list.append({
                "display"   : sel_t,
                "leave_type": sel_l
            })
            st.rerun()

if st.session_state.leave_list:
    st.subheader("Teachers on Leave")
    for idx, e in enumerate(st.session_state.leave_list):
        x, y, z = st.columns([3, 2, 1])
        x.write(f"**{e['display']}**")
        y.write(e["leave_type"])
        if z.button("❌", key=f"del_{idx}"):
            st.session_state.leave_list.pop(idx)
            st.rerun()
else:
    st.stop()

# ═══════════════════════════════════════════════════════════
#  UI — STEP 3: Arrangement Table
# ═══════════════════════════════════════════════════════════

st.divider()
st.header("Step 3: Arrangement Table")

on_leave_names = {e["display"] for e in st.session_state.leave_list}

# ── Build period rows to arrange ─────────────────────────────
arrangement_rows = []
for entry in st.session_state.leave_list:
    teacher    = entry["display"]
    leave_type = entry["leave_type"]

    # Get from Teacher TT Summary (has teacher's own class assignments)
    # Match by display name (Teacher column is already clean)
    tt_rows = teacher_tt[(teacher_tt["Teacher"] == teacher) & (teacher_tt["Day"] == day_code)]

    if tt_rows.empty:
        arrangement_rows.append({
            "teacher": teacher, "leave_type": leave_type,
            "period_col": None, "class_subject": "⚠️ No timetable found for this day"
        })
        continue

    tt_row    = tt_rows.iloc[0]
    half_cols = get_half_cols(leave_type)
    periods   = [pc for pc in half_cols if tt_row.get(pc,"")]

    if not periods:
        arrangement_rows.append({
            "teacher": teacher, "leave_type": leave_type,
            "period_col": None, "class_subject": "No classes in this half"
        })
        continue

    for pc in periods:
        arrangement_rows.append({
            "teacher"      : teacher,
            "leave_type"   : leave_type,
            "period_col"   : pc,
            "class_subject": tt_row.get(pc,""),
        })

# ── Sort rows by period (P0 → P1 → … → P8), nulls at end ────
arrangement_rows.sort(key=lambda r: (
    PERIOD_COLS.index(r["period_col"]) if r.get("period_col") else 999
))

# ── Init per-row state ────────────────────────────────────────
# State: { mode, _locked_name, _selected_radio_idx, saved }
if ("arr_state" not in st.session_state
        or len(st.session_state.arr_state) != len(arrangement_rows)):
    st.session_state.arr_state = [
        {"mode":"Same Class Teacher",
         "_locked_name": None,
         "_selected_radio_idx": None,
         "saved": False}
        for _ in arrangement_rows
    ]
arr = st.session_state.arr_state

# ── Rebuild extra_busy from saved rows ────────────────────────
extra_busy: dict = {}   # { norm(teacher_name): set(period_cols) }
for row, state in zip(arrangement_rows, arr):
    if state["saved"] and row.get("period_col") and state.get("_locked_name"):
        key = norm(state["_locked_name"])
        extra_busy.setdefault(key, set()).add(row["period_col"])

# ── Table header ─────────────────────────────────────────────
st.markdown("""
<style>
div[data-testid="stRadio"] label { font-size: 0.85rem; }
</style>
""", unsafe_allow_html=True)

h0,h1,h2,h3 = st.columns([2, 2, 2, 5])
h0.markdown("**Teacher / Period**")
h1.markdown("**Class & Subject**")
h2.markdown("**Mode**")
h3.markdown("**Select Arranging Teacher**  *(sorted: most free → least free)*")
st.divider()

for i, row in enumerate(arrangement_rows):
    c0, c1, c2, c3 = st.columns([2, 2, 2, 5])

    # ── Col 0 ──────────────────────────────
    with c0:
        st.write(f"**{row['teacher']}**")
        if row["period_col"]:
            st.caption(f"{row['period_col']} · {which_half(row['period_col'])} · {row['leave_type']}")
        else:
            st.caption(row["leave_type"])

    # ── Col 1 ──────────────────────────────
    with c1:
        st.write(row.get("class_subject","—"))

    if not row["period_col"]:
        c2.write("—"); c3.write("—")
        st.divider()
        continue

    # ── Col 2: Mode selector ───────────────
    with c2:
        if arr[i]["saved"]:
            st.write(f"**{arr[i]['mode']}**")
        else:
            mode = st.radio(
                "Mode",
                ["Same Class Teacher","Any Teacher","Manual Entry"],
                key=f"mode_{i}",
                horizontal=False,
                label_visibility="collapsed",
                index=["Same Class Teacher","Any Teacher","Manual Entry"]
                      .index(arr[i]["mode"])
            )
            arr[i]["mode"] = mode

    # ── Col 3: Candidate list / locked display ─
    with c3:
        if arr[i]["saved"]:
            locked = arr[i]["_locked_name"] or "—"
            st.write(f"✅ **Arranged by:** {locked}")

            # Show updated workload of locked teacher
            if locked and locked != "—":
                fh = count_busy_in_half(locked, day_code, FIRST_HALF,  day_tt, unavail, extra_busy)
                sh = count_busy_in_half(locked, day_code, SECOND_HALF, day_tt, unavail, extra_busy)
                lt = row["leave_type"]
                if lt == "First Half":
                    parts = [f"TT:{fh['tt']}"]
                    if fh["unav"]: parts.append(f"Unavail:{fh['unav']}")
                    if fh["extra"]: parts.append(f"Arranged:{fh['extra']}")
                    parts.append(f"Total busy 1st Half: **{fh['total']}/{len(FIRST_HALF)}**")
                    st.caption("  |  ".join(parts))
                elif lt == "Second Half":
                    parts = [f"TT:{sh['tt']}"]
                    if sh["unav"]: parts.append(f"Unavail:{sh['unav']}")
                    if sh["extra"]: parts.append(f"Arranged:{sh['extra']}")
                    parts.append(f"Total busy 2nd Half: **{sh['total']}/{len(SECOND_HALF)}**")
                    st.caption("  |  ".join(parts))
                else:
                    fd = fh["total"] + sh["total"]
                    extra = fh["extra"] + sh["extra"]
                    label = f"1st:{fh['total']}/{len(FIRST_HALF)}  2nd:{sh['total']}/{len(SECOND_HALF)}  Full day total: **{fd}/{len(PERIOD_COLS)}**"
                    if extra: label += f"  Arranged:{extra}"
                    st.caption(label)

            if st.button("✏️ Edit", key=f"edit_{i}"):
                # Remove from extra_busy
                old = arr[i]["_locked_name"]
                if old and row["period_col"] and norm(old) in extra_busy:
                    extra_busy[norm(old)].discard(row["period_col"])
                arr[i]["saved"] = False
                arr[i]["_locked_name"] = None
                arr[i]["_selected_radio_idx"] = None
                st.rerun()

        else:
            mode = arr[i]["mode"]

            if mode == "Manual Entry":
                # Only show teachers who are FREE in this specific period
                pc_for_row = row["period_col"]
                free_manual_names = [
                    t for t in day_tt_names
                    if t not in on_leave_names
                    and t != row["teacher"]
                    and is_free_in_period(t, day_code, pc_for_row, day_tt, unavail, extra_busy)
                ]

                if not free_manual_names:
                    st.warning("⚠️ No free teacher available for this period.")
                else:
                    prev_idx = 0
                    if arr[i]["_locked_name"] and arr[i]["_locked_name"] in free_manual_names:
                        prev_idx = free_manual_names.index(arr[i]["_locked_name"])
                    manual = st.selectbox(
                        "Select Teacher", free_manual_names,
                        index=prev_idx,
                        key=f"manual_{i}",
                        label_visibility="collapsed"
                    )
                    arr[i]["_locked_name"] = manual

                    # Show their workload
                    fh = count_busy_in_half(manual, day_code, FIRST_HALF,  day_tt, unavail, extra_busy)
                    sh = count_busy_in_half(manual, day_code, SECOND_HALF, day_tt, unavail, extra_busy)
                    lt = row["leave_type"]
                    if lt == "First Half":
                        st.caption(f"TT:{fh['tt']}  Unavail:{fh['unav']}  Arranged:{fh['extra']}  Busy 1st Half: **{fh['total']}/{len(FIRST_HALF)}**")
                    elif lt == "Second Half":
                        st.caption(f"TT:{sh['tt']}  Unavail:{sh['unav']}  Arranged:{sh['extra']}  Busy 2nd Half: **{sh['total']}/{len(SECOND_HALF)}**")
                    else:
                        st.caption(f"1st Half busy:{fh['total']}/{len(FIRST_HALF)}  2nd Half busy:{sh['total']}/{len(SECOND_HALF)}")

                    if st.button("✅ Confirm", key=f"add_{i}"):
                        arr[i]["saved"] = True
                        extra_busy.setdefault(norm(manual), set()).add(row["period_col"])
                        st.rerun()

            else:
                # Same Class Teacher or Any Teacher — show ranked list with radio
                candidates, fallback_used, fallback_note = get_candidate_list(
                    mode,
                    row["teacher"],
                    row["period_col"],
                    day_code,
                    teacher_tt, class_tt, day_tt,
                    unavail, all_display,
                    on_leave_names,
                    extra_busy,
                    ct_coct
                )

                if not candidates:
                    st.warning("⚠️ No free teacher found for this period even across all teachers.")
                    st.divider()
                    continue

                if fallback_used:
                    st.info(f"ℹ️ {fallback_note or 'No free same-class teacher found — showing all free teachers instead.'}")

                # Build radio labels — show role (CO-CT / 2nd CO-CT) in brackets for P0
                radio_labels = []
                for cand in candidates:
                    role     = cand.get("role")
                    role_tag = f" ({role})" if role else ""
                    lt = row["leave_type"]
                    if lt == "First Half":
                        fh = cand["fh"]
                        load = f"1st Half: TT:{fh['tt']}"
                        if fh["unav"]: load += f" Unavail:{fh['unav']}"
                        if fh["extra"]: load += f" Arranged:{fh['extra']}"
                        load += f" | Total:{fh['total']}/{len(FIRST_HALF)}"
                    elif lt == "Second Half":
                        sh = cand["sh"]
                        load = f"2nd Half: TT:{sh['tt']}"
                        if sh["unav"]: load += f" Unavail:{sh['unav']}"
                        if sh["extra"]: load += f" Arranged:{sh['extra']}"
                        load += f" | Total:{sh['total']}/{len(SECOND_HALF)}"
                    else:
                        fd = cand["fd_total"]
                        extra_ct = cand["fh"]["extra"] + cand["sh"]["extra"]
                        load = (f"1st:{cand['fh']['total']}/{len(FIRST_HALF)}  "
                                f"2nd:{cand['sh']['total']}/{len(SECOND_HALF)}  "
                                f"Full:{fd}/{len(PERIOD_COLS)}")
                        if extra_ct: load += f"  Arranged:{extra_ct}"

                    radio_labels.append(f"{cand['name']}{role_tag}  🟢 FREE  [{load}]")

                # Determine default selection index
                prev_idx = arr[i].get("_selected_radio_idx") or 0
                if prev_idx >= len(radio_labels):
                    prev_idx = 0

                selected_idx = st.radio(
                    "Choose teacher to arrange:",
                    range(len(radio_labels)),
                    format_func=lambda x: radio_labels[x],
                    key=f"radio_{i}",
                    index=prev_idx,
                    label_visibility="collapsed"
                )
                arr[i]["_selected_radio_idx"] = selected_idx
                selected_name = candidates[selected_idx]["name"]
                arr[i]["_locked_name"] = selected_name   # track current selection

                if st.button("✅ Confirm", key=f"add_{i}"):
                    arr[i]["saved"] = True
                    extra_busy.setdefault(norm(selected_name), set()).add(row["period_col"])
                    st.rerun()

    st.divider()

# ═══════════════════════════════════════════════════════════
#  SUMMARY + DOWNLOADS
# ═══════════════════════════════════════════════════════════

saved_pairs = [(r,s) for r,s in zip(arrangement_rows, arr)
               if s["saved"] and r.get("period_col")]

if saved_pairs:
    st.subheader("📄 Saved Arrangements Summary")
    out_rows = []
    for r, s in saved_pairs:
        out_rows.append({
            "Teacher on Leave": r["teacher"],
            "Period"          : r["period_col"],
            "Half"            : which_half(r["period_col"]),
            "Class & Subject" : r["class_subject"],
            "Leave Type"      : r["leave_type"],
            "Mode"            : s["mode"],
            "Arranged By"     : s.get("_locked_name") or "—",
        })
    st.dataframe(pd.DataFrame(out_rows), use_container_width=True, hide_index=True)

    total_arrangeable = sum(1 for r in arrangement_rows if r.get("period_col"))
    total_saved       = len(saved_pairs)
    sheet_name        = sel_date.strftime("%d-%m-%Y")

    if total_saved < total_arrangeable:
        st.info(f"ℹ️ {total_saved}/{total_arrangeable} period(s) confirmed — downloads include confirmed periods only.")
    else:
        st.success("✅ All periods arranged!")

    st.subheader("📥 Download Options")
    dl1, dl2 = st.columns(2)

    with dl1:
        standalone = build_standalone(sel_date, day_code, arrangement_rows, arr)
        st.download_button(
            label="📄 Download Arrangement Only",
            data=standalone,
            file_name=f"Arrangement_{sheet_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Single sheet with today's arrangement"
        )

    with dl2:
        appended = build_appended(summary_bytes, sel_date, day_code, arrangement_rows, arr)
        st.download_button(
            label=f"📚 Download Full Report (Original + {sheet_name})",
            data=appended,
            file_name=f"Summary_Report_{sheet_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Original Summary_Report with arrangement sheet appended"
        )
